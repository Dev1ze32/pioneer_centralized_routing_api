import os
import tempfile
import openpyxl
from flask import Blueprint, send_file, after_this_request, current_app, jsonify
from routes.utils.decorators import require_superuser_or_admin
from sqlalchemy import text
from db import managed_connection
from extension import limiter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

export_bp = Blueprint('export_bp', __name__)

# BUG-07: Hard row cap. Pulling 50,000 rows + building a full xlsx workbook
# in a BytesIO buffer can consume 100-200 MB of RAM per concurrent export
# request, causing OOM crashes under load. This cap prevents that.
_ROW_CAP = 10_000

# SQL query: column order mirrors the website table left-to-right.
_EXPORT_QUERY = text("""
    SELECT
        -- ── Product Metadata ──────────────────────────────────────────
        p.inventory_id              AS "Inventory ID",
        p.revision_descr            AS "Revision Descr.",
        p.revision                  AS "Revision",
        p.quantity                  AS "Qty",
        p.product_type              AS "Product Type",
        p.fg_production_line        AS "FG Production Line",
        p.fg_production_line_code   AS "FG Production Line Code",
        p.bm_production_line        AS "BM Production Line",
        p.bm_production_line_code   AS "BM Production Line Code",
        p.notes                     AS "Notes",

        -- ── ROUTING DETAILS (per activity, matches UI columns) ────────
        a.activity_name             AS "Activities",
        a.pax                       AS "Pax",
        a.machine                   AS "Machine",
        a.time_min                  AS "Time (min)",
        a.run_time                  AS "Run Time",
        a.labor_min                 AS "Total Labor (min)",
        a.mc_min                    AS "Total MC (min)",

        -- ── ACUMATICA BOM (per activity, matches UI columns) ──────────
        a.dl_units                  AS "DL (UNITS/1 MIN)",
        a.dl                        AS "DL",
        a.voh                       AS "VOH",
        a.foh                       AS "FOH",

        -- ── Extra activity metadata ────────────────────────────────────
        a.type                      AS "Type",
        a."class"                   AS "CLASS",
        a.class_1                   AS "CLASS.1",
        a.item_id                   AS "Item ID",

        -- ── Product-level TOTALS (equivalent to the footer row) ────────
        p.total_run_time            AS "Total Run Time",
        p.total_labor_min           AS "Total Labor (min) [Sum]",
        p.total_mc_min              AS "Total MC (min) [Sum]",
        p.total_dl                  AS "Total DL [Sum]",
        p.total_voh                 AS "Total VOH [Sum]",
        p.total_foh                 AS "Total FOH [Sum]"

    FROM products p
    LEFT JOIN activities a ON p.inventory_id = a.inventory_id
    ORDER BY p.inventory_id, a.sort_order
""")

_COUNT_QUERY = text("""
    SELECT COUNT(*) AS total
    FROM products p
    LEFT JOIN activities a ON p.inventory_id = a.inventory_id
""")


@export_bp.route('/api/export', methods=['GET'])
@require_superuser_or_admin
@limiter.limit("2/minute")
def export_excel():
    """
    Export all products and their activities to an Excel file.

    Column order mirrors the website table exactly:
      Product Info → Routing Details (per activity) → BOM Details (per activity) → Product Totals

    BUG-07 FIX: Row count is checked first. If it exceeds _ROW_CAP the
    request is rejected with a 400 instead of letting the server OOM.
    The xlsx is written to a temp file on disk and streamed, so the full
    serialised file is never held in a BytesIO buffer in RAM.
    """
    try:
        with managed_connection() as conn:
            # 1. Count first — refuse if the dataset is too large.
            total_rows = conn.execute(_COUNT_QUERY).mappings().first()["total"]
            if total_rows > _ROW_CAP:
                return jsonify({
                    "error": (
                        f"Export refused: {total_rows:,} rows exceed the "
                        f"{_ROW_CAP:,}-row safety limit. "
                        "Please contact your system administrator to perform "
                        "a direct database backup instead."
                    ),
                    "total_rows": total_rows,
                    "row_cap":    _ROW_CAP,
                }), 400

            # 2. Fetch rows (safe now that we know it's within the cap).
            result   = conn.execute(_EXPORT_QUERY)
            colnames = list(result.keys())
            rows     = [tuple(row) for row in result.all()]

        # ── Build Workbook ─────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ACU Routing"

        # ── Colour palette (mirrors the website) ──────────────────────────────
        TEAL_FILL   = PatternFill("solid", fgColor="006666")   # section header teal
        YELLOW_FILL = PatternFill("solid", fgColor="FFFF99")   # routing input yellow
        BOM_FILL    = PatternFill("solid", fgColor="CCE5FF")   # bom blue
        TOTAL_FILL  = PatternFill("solid", fgColor="E2EFDA")   # totals green
        WHITE_FONT  = Font(bold=True, color="FFFFFF")
        BOLD_FONT   = Font(bold=True)

        # Section boundaries (1-indexed column numbers in the sheet)
        # Product metadata: cols 1-10
        # Routing Details:  cols 11-17
        # BOM Details:      cols 18-21
        # Extra metadata:   cols 22-25
        # Totals:           cols 26-31

        SECTION_RANGES = {
            "Product Info":      (1,  10, TEAL_FILL),
            "Routing Details":   (11, 17, YELLOW_FILL),
            "Acumatica BOM":     (18, 21, BOM_FILL),
            "Extra Info":        (22, 25, TEAL_FILL),
            "Product Totals":    (26, 31, TOTAL_FILL),
        }

        # Row 1: Section header row
        ws.append([None] * len(colnames))
        for section_name, (start, end, fill) in SECTION_RANGES.items():
            ws.merge_cells(
                start_row=1, start_column=start,
                end_row=1,   end_column=end
            )
            cell = ws.cell(row=1, column=start)
            cell.value     = section_name
            cell.fill      = fill
            cell.font      = WHITE_FONT if fill in (TEAL_FILL,) else BOLD_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Row 2: Column name headers
        ws.append(colnames)
        for col_idx, col_name in enumerate(colnames, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font      = BOLD_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            # Colour the header cell to match its section
            for _, (start, end, fill) in SECTION_RANGES.items():
                if start <= col_idx <= end:
                    cell.fill = fill
                    break

        # Row 3+: Data rows
        for row in rows:
            ws.append(list(row))

        # ── Auto-fit column widths ─────────────────────────────────────────────
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                cell.alignment = Alignment(wrap_text=True)
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

        # Freeze the first two rows (section headers + column names)
        ws.freeze_panes = "A3"

        # ── BUG-07 FIX: Stream from a temp file on disk ───────────────────────
        # Writing to a NamedTemporaryFile avoids holding the entire serialised
        # xlsx in RAM. The file is deleted automatically after the response is
        # sent (via after_this_request), preventing temp-file accumulation.
        tmp = tempfile.NamedTemporaryFile(
            suffix=".xlsx", delete=False, dir=tempfile.gettempdir()
        )
        tmp_path = tmp.name
        tmp.close()

        try:
            wb.save(tmp_path)

            @after_this_request
            def _cleanup(response, _path=tmp_path):
                try:
                    os.unlink(_path)
                except OSError:
                    pass
                return response

            return send_file(
                tmp_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='acu_routing_export.xlsx',
            )
        except Exception:
            # Clean up temp file on any error path as well
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
            raise

    except Exception as e:
        current_app.logger.error(f"Error exporting Excel: {e}")
        return jsonify({"error": str(e)}), 500
